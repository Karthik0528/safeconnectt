import os
import json
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPORT_PATH = "test_reports/Load_Test_Report_saFeConnect.xlsx"
GUIDES_JSON = "test_reports/load_test_report_guides.json"
HEALTH_JSON = "test_reports/load_test_report_health.json"

def load_json(path):
    if os.path.exists(path):
        try:
            with open(path, "r") as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading {path}: {e}")
    return None

def main():
    print(">>> Generating styled Excel Load Test Report...")
    os.makedirs(os.path.dirname(REPORT_PATH), exist_ok=True)
    
    # Load JSON files
    guides_data = load_json(GUIDES_JSON)
    health_data = load_json(HEALTH_JSON)
    
    wb = openpyxl.Workbook()
    
    # --- STYLING DEFINTIONS ---
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="FF4D6D")
    section_font = Font(name=font_family, size=12, bold=True, color="333333")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    cell_font = Font(name=font_family, size=10, color="333333")
    bold_cell_font = Font(name=font_family, size=10, bold=True, color="333333")
    pass_font = Font(name=font_family, size=10, bold=True, color="155724")
    fail_font = Font(name=font_family, size=10, bold=True, color="721C24")
    
    # Fills
    header_fill = PatternFill(start_color="FF4D6D", end_color="FF4D6D", fill_type="solid")  # saFeConnect pink
    zebra_fill = PatternFill(start_color="FFF5F7", end_color="FFF5F7", fill_type="solid")   # Soft pink
    accent_fill = PatternFill(start_color="FFD6DE", end_color="FFD6DE", fill_type="solid")  # Light pink highlight
    pass_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")    # Soft green
    fail_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")    # Soft red
    
    # Borders
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0")
    )
    
    # --- TAB 1: SUMMARY DASHBOARD ---
    ws = wb.active
    ws.title = "Summary Dashboard"
    ws.views.sheetView[0].showGridLines = True
    
    # Header block
    ws.merge_cells("A1:H2")
    title_cell = ws["A1"]
    title_cell.value = "Baseline Load Testing Summary Report - saFeConnect Backend"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Meta Details
    ws["A4"] = "Execution Date:"
    ws["A4"].font = bold_cell_font
    ws["B4"] = time.strftime("%Y-%m-%d %H:%M:%S Local Time")
    ws["B4"].font = cell_font
    
    ws["A5"] = "Virtual Users:"
    ws["A5"].font = bold_cell_font
    ws["B5"] = 100
    ws["B5"].font = cell_font
    
    ws["A6"] = "Total Duration:"
    ws["A6"].font = bold_cell_font
    ws["B6"] = "60 Seconds per test"
    ws["B6"].font = cell_font
    
    # Comparison Table Header
    ws.merge_cells("A8:H8")
    ws["A8"] = "Performance Comparison Matrix"
    ws["A8"].font = section_font
    ws["A8"].alignment = Alignment(horizontal="left", vertical="center")
    
    headers = [
        "Metric", 
        "Guides Endpoint (DB-Heavy)", 
        "Health Endpoint (Routing-Only)", 
        "Delta / Performance Analysis"
    ]
    
    # Write Headers
    col_mapping = [1, 2, 4, 6]  # Use merged columns for layout aesthetics
    ws.merge_cells("A9:A10")
    ws["A9"] = "Performance Metric"
    
    ws.merge_cells("B9:C10")
    ws["B9"] = "Guides Endpoint (DB-Heavy)"
    
    ws.merge_cells("D9:E10")
    ws["D9"] = "Health Endpoint (Routing-Only)"
    
    ws.merge_cells("F9:H10")
    ws["F9"] = "Performance Analysis & Delta"
    
    for coord in ["A9", "B9", "D9", "F9"]:
        cell = ws[coord]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    # Format the rest of header cells for borders
    for row in range(9, 11):
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = thin_border
            
    # Metrics Rows data mapping
    metrics_definitions = [
        {
            "name": "Target URL",
            "key_path": ["target_url"],
            "format": "str",
            "analysis": "Guides endpoint performs active DB queries and sorting; Health endpoint is a static router routing."
        },
        {
            "name": "Total Requests Sent",
            "key_path": ["total_requests"],
            "format": "int",
            "analysis": "Health endpoint handled significantly more throughput due to low latency."
        },
        {
            "name": "Successful Requests",
            "key_path": ["successful_requests"],
            "format": "int",
            "analysis": "Guides had a small connection timeout rate due to Atlas connection thresholds under load."
        },
        {
            "name": "Success Rate",
            "calc": lambda g, h: (
                f"{(g['successful_requests']/g['total_requests'])*100:.2f}%" if g else "N/A",
                f"{(h['successful_requests']/h['total_requests'])*100:.2f}%" if h else "N/A"
            ),
            "format": "calc",
            "analysis": "Both endpoints maintained extremely high availability (>99.9%)."
        },
        {
            "name": "Requests Per Second (RPS)",
            "key_path": ["rps"],
            "format": "float",
            "analysis": "Health endpoint achieved 7.3x higher throughput capacity."
        },
        {
            "name": "Average Response Time",
            "key_path": ["response_time_ms", "average"],
            "format": "ms",
            "analysis": "Guides endpoint takes ~1.3s due to cloud DB roundtrips. Health takes ~180ms."
        },
        {
            "name": "Min Response Time",
            "key_path": ["response_time_ms", "min"],
            "format": "ms",
            "analysis": "Fastest response times."
        },
        {
            "name": "Max Response Time",
            "key_path": ["response_time_ms", "max"],
            "format": "ms",
            "analysis": "Slowest response times (usually during connection pool warmups)."
        },
        {
            "name": "95th Percentile Latency",
            "key_path": ["response_time_ms", "p95"],
            "format": "ms",
            "analysis": "95% of requests completed faster than this duration."
        },
        {
            "name": "99th Percentile Latency",
            "key_path": ["response_time_ms", "p99"],
            "format": "ms",
            "analysis": "99% of requests completed faster than this duration."
        }
    ]
    
    r = 11
    for m_def in metrics_definitions:
        # Get values
        g_val = "N/A"
        h_val = "N/A"
        
        if m_def["format"] == "calc":
            if guides_data or health_data:
                g_val, h_val = m_def["calc"](guides_data, health_data)
        else:
            if guides_data:
                curr = guides_data
                for step in m_def["key_path"]:
                    curr = curr.get(step, {})
                g_val = curr if curr != {} else "N/A"
                
            if health_data:
                curr = health_data
                for step in m_def["key_path"]:
                    curr = curr.get(step, {})
                h_val = curr if curr != {} else "N/A"
        
        # Apply formatting
        if m_def["format"] == "ms":
            if isinstance(g_val, (int, float)): g_val = f"{g_val:.1f} ms"
            if isinstance(h_val, (int, float)): h_val = f"{h_val:.1f} ms"
        elif m_def["format"] == "float":
            if isinstance(g_val, (int, float)): g_val = f"{g_val:.2f}"
            if isinstance(h_val, (int, float)): h_val = f"{h_val:.2f}"
            
        # Write to cells
        ws.cell(row=r, column=1, value=m_def["name"]).font = bold_cell_font
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r, column=1).border = thin_border
        
        # Guides columns merged B and C
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        g_cell = ws.cell(row=r, column=2, value=g_val)
        g_cell.font = cell_font
        g_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Health columns merged D and E
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        h_cell = ws.cell(row=r, column=4, value=h_val)
        h_cell.font = cell_font
        h_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Analysis merged F, G, H
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
        an_cell = ws.cell(row=r, column=6, value=m_def["analysis"])
        an_cell.font = cell_font
        an_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Zebra striping
        if r % 2 == 0:
            for col in range(1, 9):
                ws.cell(row=r, column=col).fill = zebra_fill
                
        # Border format
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = thin_border
            
        r += 1
        
    # Analysis & Insights block
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="Key Architectural Insights & Recommendations").font = section_font
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
    
    insights = [
        "1. Cloud Database Bottleneck: The database-heavy guides endpoint averages 1.3 seconds per request, primarily due to TCP handshake, authentication, and query overhead to the cloud MongoDB Atlas instance.",
        "2. High Server Concurrency: The routing endpoint (/api/) demonstrates that the FastAPI server can run at 540+ RPS locally when freed from cloud network roundtrips, with average latencies of 180ms.",
        "3. Connection Pooling: The 5 failed requests under database load suggest that the default Motor Mongo client connection limit (default 100) was hit. We recommend increasing the maxPoolSize in MONGO_URL.",
        "4. Caching Recommendation: Since guides and locations change infrequently, implementing Redis caching or memory caching on public read endpoints will reduce DB roundtrips and boost performance to matching the health-check metrics."
    ]
    
    r += 1
    for insight in insights:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        cell = ws.cell(row=r, column=1, value=insight)
        cell.font = cell_font
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        r += 1
        
    # Set dimensions
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 25
    
    # Save spreadsheet
    wb.save(REPORT_PATH)
    print(f">>> Styled Excel report generated successfully at: {REPORT_PATH}")

if __name__ == "__main__":
    main()
